"""
Excel export service — generates formatted .xlsx files using openpyxl.
"""

import io
import uuid
from datetime import date
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter
from sqlalchemy import select, and_
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import joinedload

from app.models.transaction import Transaction
from app.models.category import Category


async def export_transactions(
    db: AsyncSession,
    user_id: uuid.UUID,
    start_date: date | None = None,
    end_date: date | None = None,
) -> tuple[io.BytesIO, str]:
    """
    Export user transactions to a formatted Excel file.
    Returns (file_bytes, filename).
    """
    # Query transactions with category info
    query = (
        select(Transaction)
        .options(joinedload(Transaction.category))
        .where(Transaction.user_id == user_id)
    )

    if start_date:
        query = query.where(Transaction.transaction_date >= start_date)
    if end_date:
        query = query.where(Transaction.transaction_date <= end_date)

    query = query.order_by(Transaction.transaction_date.desc())

    result = await db.execute(query)
    transactions = result.scalars().unique().all()

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Lançamentos"

    # ── Styles ──────────────────────────────────────────────────────
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    title_font = Font(name="Calibri", bold=True, size=14, color="1A1A2E")
    date_font = Font(name="Calibri", size=10, color="666666")

    receita_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    despesa_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")

    thin_border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    # ── Title ───────────────────────────────────────────────────────
    ws.merge_cells("A1:E1")
    ws["A1"] = "Controle Financeiro — Lançamentos"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")

    period_text = "Período: "
    if start_date and end_date:
        period_text += f"{start_date.strftime('%d/%m/%Y')} a {end_date.strftime('%d/%m/%Y')}"
    elif start_date:
        period_text += f"A partir de {start_date.strftime('%d/%m/%Y')}"
    elif end_date:
        period_text += f"Até {end_date.strftime('%d/%m/%Y')}"
    else:
        period_text += "Todos os registros"

    ws.merge_cells("A2:E2")
    ws["A2"] = period_text
    ws["A2"].font = date_font
    ws["A2"].alignment = Alignment(horizontal="center")

    # ── Headers ─────────────────────────────────────────────────────
    headers = ["Data", "Tipo", "Categoria", "Descrição", "Valor (R$)"]
    header_row = 4

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # ── Data rows ───────────────────────────────────────────────────
    for row_idx, tx in enumerate(transactions, header_row + 1):
        row_fill = receita_fill if tx.type == "RECEITA" else despesa_fill

        # Date
        cell = ws.cell(
            row=row_idx, column=1,
            value=tx.transaction_date.strftime("%d/%m/%Y")
        )
        cell.border = thin_border
        cell.fill = row_fill

        # Type
        cell = ws.cell(row=row_idx, column=2, value=tx.type)
        cell.border = thin_border
        cell.fill = row_fill
        cell.alignment = Alignment(horizontal="center")

        # Category
        cat_name = tx.category.name if tx.category else "—"
        cell = ws.cell(row=row_idx, column=3, value=cat_name)
        cell.border = thin_border
        cell.fill = row_fill

        # Description
        cell = ws.cell(row=row_idx, column=4, value=tx.description or "")
        cell.border = thin_border
        cell.fill = row_fill

        # Amount
        amount_val = float(tx.amount)
        if tx.type == "DESPESA":
            amount_val = -amount_val
        cell = ws.cell(row=row_idx, column=5, value=amount_val)
        cell.number_format = '#,##0.00'
        cell.border = thin_border
        cell.fill = row_fill
        cell.alignment = Alignment(horizontal="right")

    # ── Column widths ───────────────────────────────────────────────
    col_widths = [14, 12, 22, 35, 18]
    for idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # ── Summary row ─────────────────────────────────────────────────
    if transactions:
        summary_row = header_row + len(transactions) + 2
        total_receitas = sum(
            float(t.amount) for t in transactions if t.type == "RECEITA"
        )
        total_despesas = sum(
            float(t.amount) for t in transactions if t.type == "DESPESA"
        )
        saldo = total_receitas - total_despesas

        summary_font = Font(name="Calibri", bold=True, size=11)

        ws.cell(row=summary_row, column=4, value="Total Receitas:").font = summary_font
        ws.cell(row=summary_row, column=5, value=total_receitas).font = summary_font
        ws.cell(row=summary_row, column=5).number_format = '#,##0.00'

        ws.cell(row=summary_row + 1, column=4, value="Total Despesas:").font = summary_font
        ws.cell(row=summary_row + 1, column=5, value=-total_despesas).font = summary_font
        ws.cell(row=summary_row + 1, column=5).number_format = '#,##0.00'

        ws.cell(row=summary_row + 2, column=4, value="Saldo:").font = Font(
            name="Calibri", bold=True, size=12,
            color="2E7D32" if saldo >= 0 else "C62828"
        )
        ws.cell(row=summary_row + 2, column=5, value=saldo).font = Font(
            name="Calibri", bold=True, size=12,
            color="2E7D32" if saldo >= 0 else "C62828"
        )
        ws.cell(row=summary_row + 2, column=5).number_format = '#,##0.00'

    # ── Save to bytes ───────────────────────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"lancamentos_{start_date or 'inicio'}_{end_date or 'fim'}.xlsx"
    return output, filename
